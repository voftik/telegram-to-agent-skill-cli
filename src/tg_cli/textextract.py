"""Extract readable text from downloaded attachments.

pdf via pypdf, xlsx via openpyxl; docx/pptx are unzipped and parsed with
the standard library — no heavyweight dependencies. Images are not OCRed:
agents read image files directly with their own vision.
"""

from __future__ import annotations

import logging
import zipfile
from pathlib import Path
from xml.etree import ElementTree

log = logging.getLogger(__name__)

MAX_CHARS = 1_000_000
# Resource budgets for untrusted inputs (#30): a chat member must not be
# able to OOM the agent with a zip bomb or a gigantic member file.
MAX_MEMBER_BYTES = 50 * 1024 * 1024   # single zip member / plain file read cap
MAX_COMPRESSION_RATIO = 200           # zip-bomb heuristic
MAX_ZIP_MEMBERS = 2000
MAX_PDF_PAGES = 500
_PLAIN_SUFFIXES = {".txt", ".md", ".csv", ".tsv", ".json", ".log", ".xml", ".html", ".yaml", ".yml"}

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def extractable(path: Path | str) -> bool:
    suffix = Path(path).suffix.lower()
    return suffix in _PLAIN_SUFFIXES or suffix in {".pdf", ".docx", ".pptx", ".xlsx"}


def extract_text(path: Path | str) -> str | None:
    """Return extracted text, or None when the format has no extractor
    or extraction failed. Output is capped at MAX_CHARS."""
    path = Path(path)
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            text = _pdf(path)
        elif suffix == ".docx":
            text = _docx(path)
        elif suffix == ".pptx":
            text = _pptx(path)
        elif suffix == ".xlsx":
            text = _xlsx(path)
        elif suffix in _PLAIN_SUFFIXES:
            with open(path, "rb") as f:
                text = f.read(MAX_MEMBER_BYTES).decode("utf-8", errors="replace")
        else:
            return None
    except ExtractionRejected as e:
        log.warning("extraction rejected for %s: %s", path, e)
        raise
    except Exception as e:
        log.warning("text extraction failed for %s: %s", path, e)
        return None
    if text is None:
        return None
    text = text.strip()
    return text[:MAX_CHARS] if text else None


class ExtractionRejected(Exception):
    """Input exceeds the resource budget for untrusted files (#30)."""


def _safe_zip_member(zf: zipfile.ZipFile, name: str) -> bytes:
    info = zf.getinfo(name)
    if info.file_size > MAX_MEMBER_BYTES:
        raise ExtractionRejected(f"zip member {name} expands to {info.file_size} bytes")
    if info.compress_size and info.file_size / max(info.compress_size, 1) > MAX_COMPRESSION_RATIO:
        raise ExtractionRejected(f"zip member {name} looks like a zip bomb")
    with zf.open(name) as fh:
        return fh.read(MAX_MEMBER_BYTES + 1)


def _check_zip(zf: zipfile.ZipFile) -> None:
    infos = zf.infolist()
    if len(infos) > MAX_ZIP_MEMBERS:
        raise ExtractionRejected(f"archive has {len(infos)} members")
    total = sum(i.file_size for i in infos)
    if total > 4 * MAX_MEMBER_BYTES:
        raise ExtractionRejected(f"archive expands to {total} bytes")


def _pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    total = 0
    for i, page in enumerate(reader.pages):
        if i >= MAX_PDF_PAGES or total >= MAX_CHARS:
            parts.append("… (truncated)")
            break
        chunk = page.extract_text() or ""
        parts.append(chunk)
        total += len(chunk)
    return "\n\n".join(parts)


def _docx(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        _check_zip(zf)
        root = ElementTree.fromstring(_safe_zip_member(zf, "word/document.xml"))
    paragraphs = []
    for p in root.iter(f"{_W_NS}p"):
        runs = [t.text for t in p.iter(f"{_W_NS}t") if t.text]
        if runs:
            paragraphs.append("".join(runs))
    return "\n".join(paragraphs)


def _pptx(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        _check_zip(zf)
        slide_names = sorted(
            (n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")),
            key=lambda n: int("".join(ch for ch in n if ch.isdigit()) or 0),
        )
        parts = []
        for idx, name in enumerate(slide_names, 1):
            root = ElementTree.fromstring(_safe_zip_member(zf, name))
            texts = [t.text for t in root.iter(f"{_A_NS}t") if t.text]
            if texts:
                parts.append(f"--- slide {idx} ---\n" + "\n".join(texts))
    return "\n\n".join(parts)


def _xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    try:
        for ws in wb.worksheets:
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 2000:
                    rows.append("… (truncated at 2000 rows)")
                    break
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                parts.append(f"=== {ws.title} ===\n" + "\n".join(rows))
    finally:
        wb.close()
    return "\n\n".join(parts)
